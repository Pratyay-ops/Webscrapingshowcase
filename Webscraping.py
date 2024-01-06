import csv
from bs4 import BeautifulSoup
#from msedge.selenium_tools import Edge, EdgeOptions
import requests
import re
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from pathlib import Path
from tkinter import *
def scrape_amazon(search_term):
    options = Options()
    options.add_argument('--headless')  # Run Chrome in headless mode
    options.add_argument('--no-sandbox')  # Required for running in Linux environment
    options.add_argument('--disable-dev-shm-usage')  # Required for running in Linux environment
    options.add_argument('executable_path=/home/pratyay/Documents/Vs Code/Webscraping_showcase/chromedriver_linux64/chromedriver')
    driver = webdriver.Chrome(options=options)
    driver.get(f'https://www.amazon.in/s?k={search_term}&ref=nb_sb_noss_2')
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    results = soup.find_all('div', {'data-component-type': 's-search-result'})
    records = []
    for item in results:
        description = item.h2.a.text.strip()
        price_element = item.find('span', 'a-price')
        price = price_element.find('span', 'a-offscreen').text[1:] if price_element else 'N/A'
        rating = item.i.text if item.i else ''
        review_count_element = item.find('span', {'class': 'a-size-base s-underline-text'})
        review_count = review_count_element.text if review_count_element else ''
        url = 'https://www.amazon.in' + item.h2.a.get('href')

        record = (description, price, rating, review_count, url)
        records.append(record)
    driver.quit()
    # saving data to a CSV file
    with open('Webscraping_showcase/amazon_scraped_data.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Description', 'Price in Rupees', 'Rating', 'ReviewCount', 'Url'])
        writer.writerows(records)
    return None
def FPRICE(search_term):
    def get_url(search_text):
        search_term = search_text.replace(' ', '+')
        bare_url = 'http://www.flipkart.com/search?q=' + search_term + '&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=off&as=off'
        url = bare_url
        return url
    def get_card_data(card, i):
        # Extract data from each card (THERE ARE TWO TYPES)
        if i == 2 or i == '2':
            title = card.find('a', {"class": "s1Q9rs"}).text
            try:
                price = card.find('div', {"class": "_30jeq3 _1_WHN1"}).text
            except AttributeError:
                price = ""
            try:
                # Rating count
                ratings = card.find('div', {'class': "_3LWZlK"}).text
            except AttributeError:
                ratings = ""
            try:
                # Review count
                reviews = card.find('span', {'class': "_2_R_DZ"}).text + ' ratings'
            except AttributeError:
                reviews = ""
        elif i == 1 or i == '1':
            title = card.find('a', {"class": "s1Q9rs"}).text
            try:
                price = card.find('div', {"class": "_30jeq3"}).text
            except AttributeError:
                price = ""
            try:
                # Rating count
                ratings = card.find('div', {'class': "_3LWZlK"}).text
            except AttributeError:
                ratings = ""
            try:
                # Review count
                reviews = card.find('span', {'class': "_2_R_DZ"}).text + ' ratings'
            except AttributeError:
                reviews = ""
        return (title, price[1:], ratings, reviews)
    def extract_page_data():
        # Extract and return data from the current page
        url = get_url(search_term)
        options = Options()
        options.add_argument('--headless')  # Run Chrome in headless mode
        options.add_argument('--no-sandbox')  # Required for running in Linux environment
        options.add_argument('--disable-dev-shm-usage')  # Required for running in Linux environment
        driver = webdriver.Chrome(options=options)
        driver.get(url)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        cards1 = soup.findAll('div', {'class': '_4ddWXP'})
        cards2 = soup.findAll('div', {'class': '_2kHMtA'})
        if cards1 == []:
            cards = cards2
            i = 2
        else:
            cards = cards1
            i = 1
        data = [get_card_data(card, i) for card in cards]
        return data
    product_data = extract_page_data()
    # saving data to csv file
    with open('Webscraping_showcase/flipkartproduct.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Title', 'Price in rupees', 'Rating', 'Reviews'])
        writer.writerows(product_data)
    return None
def STOCK():
    wait_imp = 10
    excel_path = Path(r"/home/pratyay/Documents/user-agent-list-master/stocks_data.xlsx")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["CMP"]

    # Read company name from excelsheet
    print ("Step 1 --> Reading Excel-sheet, Please wait....")
    s_row     = 4
    c_list    = []
    avg_val   = []
    qnty_list = []
    while ws.cell(row = s_row, column= 2).value != None:
        c_name = ws.cell(row = s_row, column= 2).value
        val_1 = ws.cell(row = s_row, column= 4).value
        qnty = ws.cell(row = s_row, column= 5).value
        c_list.append(c_name)
        avg_val.append(val_1)
        qnty_list.append(qnty)
        s_row += 1    
    print ("Company name available in Database")
    [print('    ->',name) for name in c_list]
    time.sleep(2)
    print ('\n')
    # create a webdriver object for chrome-option and configure
    CO = webdriver.ChromeOptions()
    CO.add_experimental_option('useAutomationExtension', False)
    CO.add_argument('--ignore-certificate-errors')
    CO.add_argument('--start-maximized')
    wd = webdriver.Chrome(r'C:\Users\sukro\Downloads\user-agent-list-master\chromedriver.exe',options=CO)

    print ("Step 2 --> Opening Finance website\n")
    wd.implicitly_wait(wait_imp)
    wd.get("https://www.moneycontrol.com")
    time.sleep(5)
    print ("******************************************************************************")
    print ("                      Getting Live Stock Value !! Please wait ...\n")


    for i in range(len(c_list)):
        src = wd.find_element_by_id ("search_str")
        src.send_keys(c_list[i])
        src.send_keys(Keys.RETURN)
        wd.implicitly_wait(wait_imp)
        s_v = wd.find_element_by_xpath("//*[@id='stk_overview']/div/div/div[1]/table/tbody/tr[1]/td[2]") 
        a=s_v.text
        b=a.replace(',',"")
        ws.cell(row=4+i, column= 3, value = s_v.text)
        diff = (avg_val[i] - float(b))* qnty_list[i]
        per_diff = (diff/(avg_val[i]*qnty_list[i]))*100
        print ("{:>23} -> CMP {:<7} Current P/L->[{:>8.2f}] %P/L -> {:>6.2f}%".format(c_list[i],b, diff, per_diff))

    print ('\n')
    print ("Step 3 --> Writing Latest Price into Excel-sheet ....\n")
    time.sleep(1)
    wb.save(excel_path)

    print ("Step 4 --> Successfully Written  \n")
    print ("Step 5 --> Closing browser !\n")
    print (" ----------------------- FINISHED !! ------------------------")
    time.sleep(1)
    wd.close()
    return None
from tkinter import *
import time

def click():
    entered_text=textentry.get()
    output.delete(0.0, END)
    if entered_text in ("1",1,"PRIZE COMPARISON","prize comparison","prize","Prize","Prize comparison"):
        msg="PRIZE COMPARISON PROGRAM, enter name of product you want to search"
        def click1():
            entered_text1=textentry1.get()
            try:
                scrape_amazon(entered_text1)
                FPRICE(entered_text1)
                Label(window,text="SUCCESS,now check your program folder for amazon and flipkart prices in seperate files, HAPPY SHOPPING ",bg="black",fg="white",font="none 12 bold") .grid(row=8,column=0,sticky=E)
            except:
                pass
                Label(window,text="Sorry, Something went wrong try again later!",bg="black",fg="white",font="none 12 bold") .grid(row=8,column=0,sticky=W)
        textentry1 = Entry(window, width=30, bg="white")
        textentry1.grid(row=6,column=0,sticky=W)
        Button(window,text="SUBMIT",width=6,command=click1) .grid(row=7,column=0,sticky=W)
        
    elif entered_text in ("2",2,"STOCK CHECK","STOCK CHECK PROGRAM","stock","stocks","Stock checker"):
        msg1="STOCK CHECK PROGRAM,prerequiste is that you enter your stock data in the EXCEL file given"
        output.insert(END, msg1)
        time.sleep(25)
        msg="       now check your program folder for stocks_data.xlsx"
        STOCK()
    else:
        msg="Sorry,this is not a valid option,Please enter the correct option"
    output.insert(END, msg)

window= Tk()
window.title("WELCOME TO PRIZE COMPARISON OR STOCK CHECK PROGRAM")
window.configure(background="black") 

photo1=PhotoImage(file="my.gif")
Label(window, image=photo1,bg="black") .grid(row=0,column=0,sticky=W)

Label(window,text="what you want to use 1)PRIZE COMPARISON or 2)STOCK CHECK PROGRAM",bg="black",fg="white",font="none 12 bold") .grid(row=1,column=0,sticky=W)

textentry = Entry(window, width=20, bg="white")
textentry.grid(row=2,column=0,sticky=W)

Button(window,text="SUBMIT",width=6,command=click) .grid(row=3,column=0,sticky=W)

Label(window,text="\nYou choose:",bg="black",fg="white",font="none 12 bold") .grid(row=4,column=0,sticky=W)

output=Text(window, width=75, height=2, wrap=WORD,background="white")
output.grid(row=5,column=0,columnspan=2,sticky=W)

def close_window():
    window.destroy()
    exit()
    
Label(window,text="CLICK TO EXIT",bg="black",fg="white",font="none 12 bold") .grid(row=10,column=0,sticky=W)
Button(window,text="EXIT",width=14,command=close_window) .grid(row=11,column=0,sticky=W)


window.mainloop()